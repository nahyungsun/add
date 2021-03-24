import openpyxl
from openpyxl.cell import MergedCell
import rdflib
from rdflib.namespace import DC, DCTERMS, DOAP, FOAF, SKOS, OWL, RDF, RDFS, VOID, XMLNS, XSD

class WeaponSystem:
    def __init__(self):
        self.id = ""
        self.name = ""

    def __eq__(self, other):
        """Overrides the default implementation"""
        if isinstance(other, WeaponSystem):
            return self.id == other.id
        return False

    def __str__(self):  # 문자열화 해주는 함수 선언!
        '''문자열화 해주는 함수'''

        return "id: {}, name: {}".format(self.id, self.name)


    @staticmethod
    def from_string(name):
        obj = WeaponSystem()
        if name == "-":
            obj.id = None
            obj.name = None
            return obj

        try:
            obj.id = str(name).split(")")[0][1:]
            obj.name = str(name).split(")")[1].strip()
        except Exception as e:
            print(name)
        return obj

# 수정 필요
def extract_instances_from_text(name):
    if name is None:
        return []
    instance_list = str(name).split(",")
    return instance_list


def create_dict_from_xlsx(filename):
    wb = openpyxl.load_workbook(filename=filename)

    ws = wb['참조)무기체계분류코드']

    row_index = 0
    dict = {}
    class_level_arr = []
    for i in range(0, 3):
        class_level_arr.append(None)

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):

        row_index += 1
        if str(row[0].value) == "":
            break;

        for class_level_index in range(0, 3):
            if type(row[class_level_index]) != MergedCell:
                temp_class_level = WeaponSystem.from_string(row[class_level_index].value)
                if class_level_arr[class_level_index] is None or class_level_arr[class_level_index] != temp_class_level:
                    class_level_arr[class_level_index] = temp_class_level

        #instance_list = extract_instances(str(row[3].value))



        if class_level_arr[0].id not in dict:
            dict[class_level_arr[0].id] = {}
            dict[class_level_arr[0].id]["name"] = class_level_arr[0].name
            dict[class_level_arr[0].id]["child"] = {}
        if class_level_arr[1].id not in dict[class_level_arr[0].id]["child"]:
            dict[class_level_arr[0].id]["child"][class_level_arr[1].id] = {}
            dict[class_level_arr[0].id]["child"][class_level_arr[1].id]["name"] = class_level_arr[1].name
            dict[class_level_arr[0].id]["child"][class_level_arr[1].id]["child"] = {}
        if class_level_arr[2].id not in dict[class_level_arr[0].id]["child"][class_level_arr[1].id]["child"]:
            dict[class_level_arr[0].id]["child"][class_level_arr[1].id]["child"][class_level_arr[2].id] = {}
            dict[class_level_arr[0].id]["child"][class_level_arr[1].id]["child"][class_level_arr[2].id]['name'] = class_level_arr[2].name
            dict[class_level_arr[0].id]["child"][class_level_arr[1].id]["child"][class_level_arr[2].id]["instance"] = extract_instances_from_text(str(row[3].value))

        for class_level_index in range(0, 3):
            print(class_level_arr[class_level_index], end='')
        #print(instance_list, end='')
        print("")

    wb.close()
    return dict


xlsx_file_name = "nstep_registration.xlsx"

dict = create_dict_from_xlsx(xlsx_file_name)

print(dict)
g = rdflib.Graph()

for class_1_id in dict:
    class_1_node = rdflib.URIRef("http://add.re.kr/weapon_system/"+class_1_id)
    g.add((class_1_node, RDF.type, RDFS.Class))
    g.add((class_1_node, RDFS.label, rdflib.Literal(dict[class_1_id]["name"])))

    for class_2_id in dict[class_1_id]['child']:
        class_2_node = rdflib.URIRef("http://add.re.kr/weapon_system/" + class_2_id)
        g.add((class_2_node, RDF.type, RDFS.Class))
        g.add((class_2_node, RDFS.label, rdflib.Literal(dict[class_1_id]['child'][class_2_id]["name"])))
        g.add((class_2_node, RDFS.subClassOf, class_1_node))
        for class_3_id in dict[class_1_id]['child'][class_2_id]['child']:
            if class_3_id is None:
                continue
            class_3_node = rdflib.URIRef("http://add.re.kr/weapon_system/" + class_3_id)
            g.add((class_3_node, RDF.type, RDFS.Class))
            g.add((class_3_node, RDFS.label, rdflib.Literal(dict[class_1_id]['child'][class_2_id]["child"][class_3_id]["name"])))
            g.add((class_3_node, RDFS.subClassOf, class_2_node))

            instance_id = 0
            for inst_str in dict[class_1_id]['child'][class_2_id]['child'][class_3_id]['instance']:
                class_3_instance = rdflib.URIRef("http://add.re.kr/weapon_system/" + class_3_id+"_"+str(instance_id))
                instance_id += 1
                g.add((class_3_instance, RDF.type, OWL.NamedIndividual))
                g.add((class_3_instance, RDF.type, class_3_node))
                g.add((class_3_instance, RDFS.label, rdflib.Literal(str(inst_str).strip())))



for s, p, o in g:
    print(s, p, o)

g.serialize(destination='output.ttl', format='turtle')
